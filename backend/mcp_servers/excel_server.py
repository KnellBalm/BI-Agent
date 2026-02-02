#!/usr/bin/env python3
"""
Excel MCP Server

MCP 서버로 Excel 파일을 읽고 쓰기를 수행합니다.
"""
import json
import os
import sys
from pathlib import Path
from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from .lifecycle import setup_signal_handlers, log_startup

mcp = FastMCP("excel-mcp-server")


def parse_range(range_str: str) -> tuple:
    """Parse Excel range string like 'A1:D10' into (start_row, start_col, end_row, end_col)"""
    if ':' not in range_str:
        raise ValueError(f"Invalid range format: {range_str}. Expected format: 'A1:D10'")

    start, end = range_str.split(':')

    # Parse start cell (e.g., 'A1' -> col='A', row=1)
    start_col_str = ''.join(c for c in start if c.isalpha())
    start_row = int(''.join(c for c in start if c.isdigit()))
    start_col = column_index_from_string(start_col_str)

    # Parse end cell
    end_col_str = ''.join(c for c in end if c.isalpha())
    end_row = int(''.join(c for c in end if c.isdigit()))
    end_col = column_index_from_string(end_col_str)

    return (start_row, start_col, end_row, end_col)


@mcp.tool()
def read_excel(file_path: str, sheet_name: str = None, range: str = None) -> str:
    """Read data from an Excel file

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (optional, defaults to first sheet)
        range: Cell range to read (e.g., 'A1:D10', optional)

    Returns:
        JSON with sheet name, rows data, and row count
    """
    # Validate file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    # Load workbook
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)

    # Select sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    # Read data
    if range:
        start_row, start_col, end_row, end_col = parse_range(range)

        # Get headers from first row of range
        headers = []
        for col in range(start_col, end_col + 1):
            cell_value = ws.cell(row=start_row, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"Column{col}")

        # Read data rows
        rows = []
        for row in range(start_row + 1, end_row + 1):
            row_data = {}
            for col_idx, col in enumerate(range(start_col, end_col + 1)):
                cell_value = ws.cell(row=row, column=col).value
                row_data[headers[col_idx]] = cell_value
            rows.append(row_data)
    else:
        # Read all data - first row as headers
        data = list(ws.iter_rows(values_only=True))
        if not data:
            rows = []
        else:
            headers = [str(h) if h is not None else f"Column{i}" for i, h in enumerate(data[0])]
            rows = [dict(zip(headers, row)) for row in data[1:]]

    wb.close()

    return json.dumps({
        "sheet": sheet_name,
        "rows": rows,
        "rowCount": len(rows)
    }, indent=2, default=str)


@mcp.tool()
def list_sheets(file_path: str) -> str:
    """List all sheet names in an Excel file

    Args:
        file_path: Path to the Excel file

    Returns:
        JSON array of sheet names
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(filename=file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    return json.dumps(sheet_names, indent=2)


@mcp.tool()
def write_excel(file_path: str, data: list, sheet_name: str = "Sheet1") -> str:
    """Write data to an Excel file

    Args:
        file_path: Path to save the Excel file
        data: Array of objects to write as rows (first row keys become headers)
        sheet_name: Name of the sheet (optional, defaults to 'Sheet1')

    Returns:
        JSON with success message, file path, and row count
    """
    # Validate data
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("Data must be a non-empty array")

    # Create directory if it doesn't exist
    dir_path = os.path.dirname(file_path)
    if dir_path:
        Path(dir_path).mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Get headers from first row's keys
    if isinstance(data[0], dict):
        headers = list(data[0].keys())
    else:
        raise ValueError("Data items must be objects/dictionaries")

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save file
    wb.save(file_path)
    wb.close()

    return json.dumps({
        "message": "Excel file written successfully",
        "file_path": file_path,
        "rowCount": len(data)
    }, indent=2)


if __name__ == "__main__":
    setup_signal_handlers()
    log_startup("Excel MCP Server")
    mcp.run(transport="stdio")
